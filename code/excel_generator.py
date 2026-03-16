import json
from pathlib import Path

from openpyxl import Workbook


SKIP_DIRS = {"to_translate", "translated", "to_translate_old", "translated_old"}

COLUMNS = [
    "File Name",
    "Title",
    "Channel Name",
    "Publish Date",
    "View Count",
    "Like Count",
    "Duration",
    "Original Audio Language",
    "Video ID",
    "URL",
    "Description",
    "Ingredients Detected",
]


def safe_sheet_name(name: str) -> str:
    invalid_chars = {"\\", "/", "*", "[", "]", ":", "?"}
    cleaned = "".join("_" if ch in invalid_chars else ch for ch in name)
    return cleaned[:31] if cleaned else "Sheet"


def ingredients_to_text(value) -> str:
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    if value is None:
        return ""
    return str(value)


def metadata_row(file_name: str, data: dict) -> list:
    metadata = data.get("metadata", {}) if isinstance(data, dict) else {}

    return [
        file_name,
        metadata.get("title", ""),
        metadata.get("channel_name", ""),
        metadata.get("publish_date", ""),
        metadata.get("view_count", ""),
        metadata.get("like_count", ""),
        metadata.get("duration", ""),
        metadata.get("original_audio_language", ""),
        metadata.get("video_id", ""),
        metadata.get("url", ""),
        metadata.get("description", ""),
        ingredients_to_text(metadata.get("ingredients_detected", [])),
    ]


def main() -> None:
    root_dir = Path(__file__).resolve().parent.parent
    data_dir = root_dir / "data"
    output_file = data_dir / "channel_metadata.xlsx"

    if not data_dir.exists():
        print(f"Data folder not found: {data_dir}")
        return

    channel_dirs = [
        d for d in data_dir.iterdir()
        if d.is_dir() and d.name not in SKIP_DIRS
    ]
    channel_dirs.sort(key=lambda path: path.name.lower())

    if not channel_dirs:
        print("No channel folders found inside data/.")
        return

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    total_rows = 0

    for channel_dir in channel_dirs:
        sheet = workbook.create_sheet(title=safe_sheet_name(channel_dir.name))
        sheet.append(COLUMNS)

        for json_file in sorted(channel_dir.glob("*.json"), key=lambda p: p.name.lower()):
            try:
                with open(json_file, "r", encoding="utf-8") as handle:
                    data = json.load(handle)
            except Exception as error:
                print(f"Skipping unreadable file {json_file}: {error}")
                continue

            sheet.append(metadata_row(json_file.name, data))
            total_rows += 1

    workbook.save(output_file)

    print(f"Excel generated: {output_file}")
    print(f"Sheets created: {len(workbook.sheetnames)}")
    print(f"Total JSON rows written: {total_rows}")


if __name__ == "__main__":
    main()
